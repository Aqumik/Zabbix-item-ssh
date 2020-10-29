# @aqumik
# -*- coding=utf-8 -*-
import json
import urllib2
from email.header import Header
import openpyxl
import time
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
# from ip2Region import Ip2Region
import struct, io, socket, sys
import os


#更换数据库路径，在服务器中需要用到，否则执行命令会出错
os.chdir('D:\pythonProject2.7\okk')
db_file = 'ip2region.db'


# 主要的url，url2为获取ip地址所属地的API
url = "http://xx.xx.xx.xx/zabbix/api_jsonrpc.php"       #***** Your Zabbix server address *****
header = {"Content-Type": "application/json"}
url2 = "http://ip-api.com/json/"
time = time.strftime("%Y%m%d%H%M", time.localtime())  # 报表名，时间开头




# 定义好的类，用于检索ip库
class Ip2Region(object):

    def __init__(self, dbfile):
        self.__INDEX_BLOCK_LENGTH = 12
        self.__TOTAL_HEADER_LENGTH = 8192
        self.__f = None
        self.__headerSip = []
        self.__headerPtr = []
        self.__headerLen = 0
        self.__indexSPtr = 0
        self.__indexLPtr = 0
        self.__indexCount = 0
        self.__dbBinStr = ''
        self.initDatabase(dbfile)

    def memorySearch(self, ip):
        """
        " memory search method
        " param: ip
        """
        if not ip.isdigit(): ip = self.ip2long(ip)

        if self.__dbBinStr == '':
            self.__dbBinStr = self.__f.read()  # read all the contents in file
            self.__indexSPtr = self.getLong(self.__dbBinStr, 0)
            self.__indexLPtr = self.getLong(self.__dbBinStr, 4)
            self.__indexCount = int((self.__indexLPtr - self.__indexSPtr) / self.__INDEX_BLOCK_LENGTH) + 1

        l, h, dataPtr = (0, self.__indexCount, 0)
        while l <= h:
            m = int((l + h) >> 1)
            p = self.__indexSPtr + m * self.__INDEX_BLOCK_LENGTH
            sip = self.getLong(self.__dbBinStr, p)

            if ip < sip:
                h = m - 1
            else:
                eip = self.getLong(self.__dbBinStr, p + 4)
                if ip > eip:
                    l = m + 1;
                else:
                    dataPtr = self.getLong(self.__dbBinStr, p + 8)
                    break

        if dataPtr == 0: raise Exception("Data pointer not found")

        return self.returnData(dataPtr)

    def binarySearch(self, ip):
        """
        " binary search method
        " param: ip
        """
        if not ip.isdigit(): ip = self.ip2long(ip)

        if self.__indexCount == 0:
            self.__f.seek(0)
            superBlock = self.__f.read(8)
            self.__indexSPtr = self.getLong(superBlock, 0)
            self.__indexLPtr = self.getLong(superBlock, 4)
            self.__indexCount = int((self.__indexLPtr - self.__indexSPtr) / self.__INDEX_BLOCK_LENGTH) + 1

        l, h, dataPtr = (0, self.__indexCount, 0)
        while l <= h:
            m = int((l + h) >> 1)
            p = m * self.__INDEX_BLOCK_LENGTH

            self.__f.seek(self.__indexSPtr + p)
            buffer = self.__f.read(self.__INDEX_BLOCK_LENGTH)
            sip = self.getLong(buffer, 0)
            if ip < sip:
                h = m - 1
            else:
                eip = self.getLong(buffer, 4)
                if ip > eip:
                    l = m + 1
                else:
                    dataPtr = self.getLong(buffer, 8)
                    break

        if dataPtr == 0: raise Exception("Data pointer not found")

        return self.returnData(dataPtr)

    def btreeSearch(self, ip):
        """
        " b-tree search method
        " param: ip
        """
        if not ip.isdigit(): ip = self.ip2long(ip)

        if len(self.__headerSip) < 1:
            headerLen = 0
            # pass the super block
            self.__f.seek(8)
            # read the header block
            b = self.__f.read(self.__TOTAL_HEADER_LENGTH)
            # parse the header block
            for i in range(0, len(b), 8):
                sip = self.getLong(b, i)
                ptr = self.getLong(b, i + 4)
                if ptr == 0:
                    break
                self.__headerSip.append(sip)
                self.__headerPtr.append(ptr)
                headerLen += 1
            self.__headerLen = headerLen

        l, h, sptr, eptr = (0, self.__headerLen, 0, 0)
        while l <= h:
            m = int((l + h) >> 1)

            if ip == self.__headerSip[m]:
                if m > 0:
                    sptr = self.__headerPtr[m - 1]
                    eptr = self.__headerPtr[m]
                else:
                    sptr = self.__headerPtr[m]
                    eptr = self.__headerPtr[m + 1]
                break

            if ip < self.__headerSip[m]:
                if m == 0:
                    sptr = self.__headerPtr[m]
                    eptr = self.__headerPtr[m + 1]
                    break
                elif ip > self.__headerSip[m - 1]:
                    sptr = self.__headerPtr[m - 1]
                    eptr = self.__headerPtr[m]
                    break
                h = m - 1
            else:
                if m == self.__headerLen - 1:
                    sptr = self.__headerPtr[m - 1]
                    eptr = self.__headerPtr[m]
                    break
                elif ip <= self.__headerSip[m + 1]:
                    sptr = self.__headerPtr[m]
                    eptr = self.__headerPtr[m + 1]
                    break
                l = m + 1

        if sptr == 0: raise Exception("Index pointer not found")

        indexLen = eptr - sptr
        self.__f.seek(sptr)
        index = self.__f.read(indexLen + self.__INDEX_BLOCK_LENGTH)

        l, h, dataPrt = (0, int(indexLen / self.__INDEX_BLOCK_LENGTH), 0)
        while l <= h:
            m = int((l + h) >> 1)
            offset = int(m * self.__INDEX_BLOCK_LENGTH)
            sip = self.getLong(index, offset)

            if ip < sip:
                h = m - 1
            else:
                eip = self.getLong(index, offset + 4)
                if ip > eip:
                    l = m + 1;
                else:
                    dataPrt = self.getLong(index, offset + 8)
                    break

        if dataPrt == 0: raise Exception("Data pointer not found")

        return self.returnData(dataPrt)

    def initDatabase(self, dbfile):
        """
        " initialize the database for search
        " param: dbFile
        """
        try:
            self.__f = io.open(dbfile, "rb")
        except IOError as e:
            print("[Error]: %s" % e)
            sys.exit()

    def returnData(self, dataPtr):
        """
        " get ip data from db file by data start ptr
        " param: dsptr
        """
        dataLen = (dataPtr >> 24) & 0xFF
        dataPtr = dataPtr & 0x00FFFFFF

        self.__f.seek(dataPtr)
        data = self.__f.read(dataLen)

        return {
            "city_id": self.getLong(data, 0),
            "region": data[4:]
        }

    def ip2long(self, ip):
        _ip = socket.inet_aton(ip)
        return struct.unpack("!L", _ip)[0]

    def isip(self, ip):
        p = ip.split(".")

        if len(p) != 4: return False
        for pp in p:
            if not pp.isdigit(): return False
            if len(pp) > 3: return False
            if int(pp) > 255: return False

        return True

    def getLong(self, b, offset):
        if len(b[offset:offset + 4]) == 4:
            return struct.unpack('I', b[offset:offset + 4])[0]
        return 0

    def close(self):
        if self.__f != None:
            self.__f.close()

        self.__dbBinStr = None
        self.__headerPtr = None
        self.__headerSip = None



#邮箱
def sendMail(mailto,subject,body,format='plain'):
    # if isinstance(body,unicode):
    #     body = str(body)
    me= ("%s<"+fromMail+">") % (Header("杜伟志",'utf-8'),)
    textApart = MIMEText(body, format, 'utf-8')
    msg = MIMEMultipart()
    # msg = MIMEText(body,format,'utf-8')
    # if not isinstance(subject,unicode):
    #     subject = unicode(subject)
    msg['Subject'] = subject
    msg['From'] = me
    msg['To'] = Header('运维部','utf-8')
    msg["Accept-Language"]="zh-CN"
    msg["Accept-Charset"]="ISO-8859-1,utf-8"

    xlsFile = time + '.xlsx'
    xlsApart = MIMEApplication(open(xlsFile, 'rb').read())
    xlsApart.add_header('Content-Disposition', 'attachment', filename=xlsFile)

    msg.attach(xlsApart)
    msg.attach(textApart)


    try:
        s = smtplib.SMTP_SSL('smtp.qiye.163.com', 994)
        # s.connect(host)
        s.login(user,password)
        s.sendmail(me, mailto, msg.as_string())
        s.close()
        print '邮件发送成功 '
        return True
    except Exception, e:
        print str(e)
        return False

# 关闭连接
def user_logout(authid):
    data = json.dumps(
        {
            "jsonrpc": "2.0",
            "method": "user.logout",
            "params": [],
            "id": 0,
            "auth": authid
        }
    )
    request = urllib2.Request(url, data)
    for key in header:
        request.add_header(key, header[key])
    result = urllib2.urlopen(request)
    response = json.loads(result.read())
    print response
    result.close


# 得到zabbix api
data = json.dumps(
    {
        "jsonrpc": "2.0",
        "method": "user.login",
        "params": {
            "user": "Admin",
            "password": "1qaz@WSXcde3"
        },
        "id": 0
    })
requestz = urllib2.Request(url, data)

for key in header:
    requestz.add_header(key, header[key])

try:
    resultz = urllib2.urlopen(requestz)
# result = urllib2.urlopen(request)
except urllib2.URLError as l:
    print "认证失败，请检查你的用户名和密码", l.code
else:
    responsez = json.loads(resultz.read())
    # print response1['result']
    authid = responsez['result']
    resultz.close()
    # print authid

# 取得hostid的JSON格式，划为字符串形式
data = json.dumps(
    {
        "jsonrpc": "2.0",
        "method": "host.get",
        "params": {
            "output": [
                "hostid",
                "host"
            ]
        },
        # 此处authid可以继续优化，因为每次获取都会得到不同的id，优化可以针对是否可关闭，或者每次获取一个authid
        "auth": authid,
        "id": 1,
    })

# 报表处理，可以继续对格式进行优化
book = openpyxl.Workbook()
sheet = book.create_sheet('Sheet1', 0)
proj = ['主机', 'IP', '地点', '次数']
for i in range(len(proj)):
    sheet.cell(1, i + 1, proj[i])
# 调节xlsx中列宽，便于查看，A列：主机名 B列：IP  C列：地点   D列：次数
sheet.column_dimensions['A'].width = 19.75
sheet.column_dimensions['B'].width = 19.75
sheet.column_dimensions['C'].width = 39

# 创建请求对象，request1为请求hostid
request1 = urllib2.Request(url, data)
for key in header:
    request1.add_header(key, header[key])

try:
    result1 = urllib2.urlopen(request1)
except urllib2.URLError as e:
    if hasattr(e, 'reason'):
        print '无法连接服务器'
        print 'Reason: ', e.reason
    elif hasattr(e, 'code'):
        print '服务器无法响应请求'
        print 'Error code: ', e.code
else:
    # 请求返回数据，格式为前缀带 ‘u’的unicode字符串
    response1 = json.loads(result1.read())
    result1.close()

    # print  response1

    # 遍历得到的host和hostid，再引入到item.get,获取下一步监控数据，若没数据此为转跳点
    # count计数，用于写入xlsx的行数计算
    # host_str:主机名称   hostid_str：主机id
    count = 0
    for hostdd in response1['result']:
        host_str = str(hostdd['host'])
        hostid_str = str(hostdd['hostid'])
        # print "主机名：", host_str, "hostid：", hostid_str

        # 根据hostid判断item，data2：通过上面循环得到的hostid，替换到data2的json中获得监控值
        data2 = json.dumps(
            {
                "jsonrpc": "2.0",
                "method": "item.get",
                "params": {
                    "output": "extend",
                    "hostids": hostid_str,
                    "search": {
                        "key_": "custom.ssh.failed.num"
                    },
                    "sortfield": "name"
                },
                "auth": authid,
                "id": 1,
            })

        # request2,result2都是用作”item.get“的请求和处理数据
        request2 = urllib2.Request(url, data2)
        for key in header:
            request2.add_header(key, header[key])
        try:
            result2 = urllib2.urlopen(request2)
        except urllib2.URLError as e:
            if hasattr(e, 'reason'):
                print '无法连接服务器'
                print 'Reason: ', e.reason
            elif hasattr(e, 'code'):
                print '服务器无法响应请求'
                print 'Error code: ', e.code
        else:
            # 化为字典
            response2 = json.loads(result2.read())
            result2.close()
            # print count
            # print "尝试登陆IP数量: ", len(response1['result'])

            # 此处输出[]
            # print '----'
            # print response2['result']
            # print '----'

            # 下面判断是用于判断监控主机是否有登陆失败的监控项，有，则继续循环写入xlsx报表，若没有数据，则跳出到下一台主机。
            if response2['result'] == []:
                continue
                # print("没有数据")
            else:
                # 筛选遍历出JSON中的'result'段数据,以字典形式存在
                for get_result in response2['result']:
                    # filt参数处理得到数据，得到的数据格式是 [ip],要把框内过滤，后续此处可以进行语法优化提高代码速度
                    filt = format(get_result['key_'])
                    filt_1 = filt.find('[') + 1
                    filt_2 = filt.find(']')
                    # key_ip，last_value分别对应监控的IP值，尝试登陆次数
                    key_ip = get_result['key_'][filt_1:filt_2]
                    last_value = int(get_result['lastvalue'])

                    # 输出测试
                    # print(key_ip, last_value)

                    count += 1

                    # 调用离线ip库 ip2region
                    searcher = Ip2Region(db_file)
                    if searcher.isip(key_ip):
                        # 查找算法有三种，此处用的是btree
                        data = searcher.btreeSearch(key_ip)["region"]
                    # data = search.binarySearch(key_ip)    #二进制查找
                    # data = search.memorySearch(key_ip)    #内存查找
                    else:
                        data = '错误数据'
                    #
                    # sheet.cell(count + 1, 3, data["region"].decode('utf-8'))
                    sheet.cell(count + 1, 3, data.decode('utf-8'))

                    # # 根据监控得到的IP地址，调用查询IP归属地的API接口
                    # url3 = url2 + key_ip
                    # time.sleep(1)
                    # ip_data = urllib2.urlopen(url3)
                    # # s = ip_data.read()
                    # # t = json.loads(s)
                    # ipcheck = json.loads(ip_data.read())
                    #
                    # # 根据API接口返回的数据进行测试，每个接口返回接口会有所不同，处理代码也不一样,后续代码优化点
                    # if ipcheck['status'] == 'success':
                    #     ipcheck_1 = ipcheck['country'], ipcheck['regionName'], ipcheck['city']
                    #     ipcheck_2 = ipcheck_1[0] + ' , ' + ipcheck_1[1] + ' , ' + ipcheck_1[2]
                    #     # 先把归属地写入到xlsx
                    #     sheet.cell(count + 1, 3, ipcheck_2)
                    # # print s
                    # else:
                    #     ipcheck_2 = ipcheck['message']
                    #     # 对于查询地址失败的处理
                    #     sheet.cell(count + 1, 3, ipcheck_2)

                    # 测试计算次数
                    # print count
                    # 写入主机、IP、攻击次数
                    sheet.cell(count + 1, 1, host_str)
                    # 测试实际效果
                    # print host_str
                    sheet.cell(count + 1, 2, key_ip)
                    # sheet.cell(count + 1, 3, ac)
                    sheet.cell(count + 1, 4, last_value)

book.save(time + '.xlsx')

if __name__ == '__main__':

    user_logout(authid)
    host = 'example@xxx.com'
    user = 'example@xxx.com'
    fromMail = 'example@xxx.com'                 #mail from
    body = '各个服务器登陆失败次数汇总报表'
    subjest = 'Zabbix服务器登陆情况报表'
    password = '*******************'             #********************Your email Password！！！！**********************
    mailto = 'example@xxx.com'                   # mail to 
    sendMail(mailto, subjest, body)       

